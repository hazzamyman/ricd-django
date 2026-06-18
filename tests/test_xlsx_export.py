"""XLSX exports — single-sheet Work Items + combined multi-sheet workbook."""
import pytest
from decimal import Decimal
from io import BytesIO
from django.urls import reverse
from tests.fixtures import make_bfa


@pytest.mark.django_db
def test_work_items_xlsx(admin_client, project, work_type):
    import openpyxl
    from apps.core.models import Work

    work_type.category = 'RESIDENTIAL'
    work_type.save()
    Work.objects.create(project=project, work_type=work_type, quantity=2,
                        estimated_cost=Decimal('300000'), is_notional_cost=False,
                        actual_cost=Decimal('300000'))
    make_bfa(project, funding_amount=600000, status='APPROVED')

    resp = admin_client.get(reverse('ui:work_items_export'), {'format': 'xlsx'})
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp['Content-Type']
    assert resp.content[:2] == b'PK'  # xlsx is a zip

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert 'Work Items' in wb.sheetnames
    ws = wb['Work Items']
    assert ws['A1'].value == 'Council (LGA)'
    assert ws.freeze_panes == 'A2'          # frozen top row
    assert ws['A1'].font.bold is True        # bold header
    assert ws.auto_filter.ref is not None    # auto-filter
    assert project.name in [c.value for c in ws['C']]


@pytest.mark.django_db
def test_reports_workbook_multisheet(admin_client, project, work_type):
    import openpyxl
    from apps.core.models import Work

    work_type.category = 'RESIDENTIAL'
    work_type.save()
    Work.objects.create(project=project, work_type=work_type, quantity=1,
                        estimated_cost=Decimal('100000'), is_notional_cost=False,
                        actual_cost=Decimal('100000'))

    resp = admin_client.get(reverse('ui:reports_workbook_export'))
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    for name in ['About', 'Work Items', 'Overall', 'Land', 'Dwellings', 'Cashflow (Monthly)']:
        assert name in wb.sheetnames, f"missing sheet {name}"
