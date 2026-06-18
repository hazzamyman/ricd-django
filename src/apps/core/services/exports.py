"""
Spreadsheet export helpers (XLSX via openpyxl, shared row builders for CSV/XLSX).

Produces a multi-sheet workbook for sharing:
  * "Work Items"            — flat dump, one row per Work item.
  * one sheet per Analytics category (Overall / Land / Dwellings / …).
  * "Cashflow (Monthly)"    — per-Program × month forecast/released (long format).

Each sheet gets a bold header row, a frozen top row, an auto-filter, sensible
column widths, and number formatting on numeric cells.
"""
import datetime
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum
from django.http import HttpResponse

from apps.core.services.analytics import build_aggregate_outputs, CATS, CAT_LABEL
from apps.core.services.cashflow import build_program_monthly_cashflow

XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


# ── shared row builders ──────────────────────────────────────────────

WORK_ITEM_HEADERS = [
    'Council (LGA)', 'Region', 'Project', 'Project State', 'Financial Year', 'Program',
    'Address', 'Work Type', 'Category', 'Bedrooms', 'Quantity', 'Work Status',
    'Estimated Cost', 'Actual Cost', 'Effective Cost',
    'Project Approved Budget (BFA)', 'Project Expended (Released)',
    'Steps Total', 'Steps Completed', 'All Steps Complete', 'Last Step Completed',
    'Actual Start', 'Practical Completion (Actual)', 'Forecast PC', 'Handover (Actual)',
    'Contractor', 'Costs Finalised',
]


def work_items_rows(council=None, include_archived=False):
    """Return (headers, rows) for the all-work-items dump, with native values
    (Decimal / date / int / str) so both CSV and XLSX can consume them."""
    from apps.core.models import Work, BriefFinancialApprovalItem, PaymentAllocation

    works = (
        Work.objects
        .select_related('project__council', 'project__program', 'work_type',
                        'address', 'contractor')
        .prefetch_related('steps')
        .order_by('project__council__name', 'project__name', 'work_type__name', 'id')
    )
    if council:
        works = works.filter(project__council_id=council)
    if not include_archived:
        works = works.filter(project__is_archived=False)

    approved = dict(
        BriefFinancialApprovalItem.objects.filter(bfa__status='APPROVED')
        .values('project_id').annotate(t=Sum('funding_amount'))
        .values_list('project_id', 't')
    )
    expended = dict(
        PaymentAllocation.objects.values('payment__project_id')
        .annotate(t=Sum('amount')).values_list('payment__project_id', 't')
    )

    rows = []
    for wk in works:
        p = wk.project
        steps = [s for s in wk.steps.all() if s.is_active]
        done = [s for s in steps if s.completed or s.actual_completion_date]
        last_done = max((s.actual_completion_date for s in done if s.actual_completion_date),
                        default=None)
        wt = wk.work_type
        rows.append([
            p.council.name if p.council_id else '',
            (p.council.region if (p.council_id and p.council.region) else ''),
            p.name, p.get_state_display(), p.financial_year or '',
            p.program.name if p.program_id else '',
            str(wk.address) if wk.address_id else '',
            wt.name if wt else (wk.work_type_other or 'Other'),
            wt.get_category_display() if wt else '',
            int(wk.bedrooms or 0), int(wk.quantity or 0), wk.get_status_display(),
            wk.estimated_cost, wk.actual_cost, wk.total_effective_cost,
            approved.get(p.pk), expended.get(p.pk),
            len(steps), len(done),
            ('Yes' if (steps and len(done) == len(steps)) else 'No'),
            last_done,
            wk.actual_start_date, wk.practical_completion_date,
            wk.forecast_practical_completion_date, wk.handover_date,
            str(wk.contractor) if wk.contractor_id else '',
            ('Yes' if wk.costs_finalised else 'No'),
        ])
    return WORK_ITEM_HEADERS, rows


def analytics_sheets(region=None):
    """Yield (sheet_title, headers, rows) for each Analytics category."""
    data = build_aggregate_outputs(region=region)
    out = []
    for cat in (['overall'] + CATS):
        cd = data['data'].get(cat, {})
        progs = cd.get('programs', [])
        show_da = cat not in ('overall', 'dwellings')
        headers = ['LGA', 'Region', 'Potential', 'In pipeline', 'Funded·NC', 'Commenced',
                   'Under constr.', 'Completed', 'Funded Yield']
        headers += [f"{p['short']} (approved)" for p in progs]
        headers += ['Total Approved', 'Paid to Council']
        if show_da:
            headers += ['DA appr.', 'DA subm.', 'DA none', 'Surplus/Deficit']
        headers += ['Avg cost / unit']

        def _row(r):
            row = [r.get('council', ''), r.get('region', ''),
                   int(r.get('potential', 0)), int(r.get('inPipeline', 0)),
                   int(r.get('fundedNotCommenced', 0)), int(r.get('commenced', 0)),
                   int(r.get('underConstruction', 0)), int(r.get('completed', 0)),
                   int(r.get('fundedYield', 0))]
            funding = r.get('funding', {})
            row += [round(funding.get(p['id'], 0)) for p in progs]
            row += [round(r.get('totalApproved', 0)), round(r.get('paid', 0))]
            if show_da:
                row += [int(r.get('daApproved', 0)), int(r.get('daSubmitted', 0)),
                        int(r.get('daNotStarted', 0)), int(r.get('surplus', 0))]
            row += [round(r.get('avgCost', 0))]
            return row

        rows = [_row(r) for r in cd.get('rows', [])]
        if cd.get('rows'):
            rows.append(_row(cd.get('totals', {})))  # All-LGAs total row
        out.append((CAT_LABEL[cat], headers, rows))
    return out


def cashflow_monthly_sheet(start=None, months=24):
    """Return (title, headers, rows) for the monthly cashflow (long format)."""
    data = build_program_monthly_cashflow(start=start, months=months)
    headers = ['Program', 'CC', 'GL', 'Month', 'Forecast', 'Released']
    progs = {p['id']: p for p in data['programs']}
    sy, sm = (int(x) for x in data['start'].split('-'))
    keys = []
    y, m = sy, sm
    for _ in range(data['months']):
        keys.append(f"{y}-{m:02d}")
        m += 1
        if m > 12:
            m = 1; y += 1
    rows = []
    for pid, p in progs.items():
        for k in keys:
            c = data['cells'].get(f"{pid}|{k}")
            if not c:
                continue
            f = round(c.get('forecast') or 0)
            r = c.get('released')
            if f or r is not None:
                rows.append([p['name'], p.get('cc', ''), p.get('gl', ''), k,
                             f, (round(r) if r is not None else None)])
    return ('Cashflow (Monthly)', headers, rows)


# ── workbook assembly ────────────────────────────────────────────────

def _style_ws(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='09549F', end_color='09549F', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        maxlen = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            v = cell.value
            if row_idx > 1 and isinstance(v, (Decimal, float)):
                cell.number_format = '#,##0'
            s = '' if v is None else str(v)
            if len(s) > maxlen:
                maxlen = len(s)
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 48)


def build_workbook(sheets, about=None):
    """sheets = list of (title, headers, rows). Returns an openpyxl Workbook."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    if about:
        ws = wb.create_sheet('About')
        ws.append(['RICD — Reports workbook'])
        ws.append(['Generated', datetime.datetime.now().strftime('%Y-%m-%d %H:%M')])
        for k, v in about.items():
            ws.append([k, v])
        ws.append([])
        ws.append(['Sheets'])
        for title, _h, rows in sheets:
            ws.append([title, f"{len(rows)} rows"])
        from openpyxl.styles import Font
        ws['A1'].font = Font(bold=True, size=14)
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 40

    used = set()
    for title, headers, rows in sheets:
        safe = (title or 'Sheet')[:31]
        base = safe
        i = 2
        while safe in used:
            safe = f"{base[:28]} {i}"
            i += 1
        used.add(safe)
        ws = wb.create_sheet(safe)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        _style_ws(ws)
    return wb


def workbook_response(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type=XLSX_CONTENT_TYPE)
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
