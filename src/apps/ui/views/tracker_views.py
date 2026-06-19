"""
Monthly Tracker + Quarterly Report views (per-council).

Both follow the same pattern:
- List view (council-scoped for council users; all-councils for RICD staff)
- Detail / grid editor with row=Work, col=WorkStep (monthly) or QR Item (quarterly)
- Submit / approve lifecycle actions
- Maintenance config: CouncilTrackerConfig per council

Access rules:
- RICD staff (officer_role NOT in {COUNCIL_USER, COUNCIL_MANAGER}) can view/edit anything
- Council users see their own council only, can edit when status=DRAFT
  AND the council's CouncilTrackerConfig.council_submission_enabled is True
"""
from datetime import date

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.generic import View

from apps.core.models import (
    Council, CouncilTrackerConfig, FundingSchedule,
    MonthlyTracker, MonthlyTrackerWorkEntry,
    Project, QuarterlyReport, QuarterlyReportEntry, QuarterlyReportItem,
    Work, WorkStep,
)

COUNCIL_ROLES = frozenset({'COUNCIL_USER', 'COUNCIL_MANAGER'})
MANAGER_ROLES = frozenset({'MANAGER', 'DIRECTOR'})


def _role(user):
    return getattr(getattr(user, 'profile', None), 'officer_role', None)


def _is_council_user(user):
    return _role(user) in COUNCIL_ROLES


def _user_council(user):
    return getattr(getattr(user, 'profile', None), 'council', None)


def _is_ricd_staff(user):
    if user.is_superuser:
        return True
    role = _role(user)
    return role is not None and role not in COUNCIL_ROLES


def _active_projects_for_council(council):
    """Projects in COMMENCED or UNDER_CONSTRUCTION state for this council."""
    return Project.objects.filter(
        council=council,
        state__in=[Project.State.COMMENCED, Project.State.UNDER_CONSTRUCTION],
    )


# ===========================================================================
# CouncilTrackerConfig (maintenance UI)
# ===========================================================================

class CouncilTrackerConfigListView(LoginRequiredMixin, View):
    """List councils with their tracker config (RICD staff only)."""

    def get(self, request):
        if not _is_ricd_staff(request.user):
            messages.error(request, "Only RICD staff can manage tracker configuration.")
            return redirect('ui:dashboard')

        councils = Council.objects.order_by('name')
        rows = []
        for c in councils:
            cfg = CouncilTrackerConfig.objects.filter(council=c).first()
            rows.append({
                'council': c,
                'config': cfg,
                'enabled': cfg.council_submission_enabled if cfg else False,
                'due_day': cfg.submission_due_day if cfg else 8,
            })
        return render(request, 'tracker/config_list.html', {'rows': rows})


class CouncilTrackerConfigUpdateView(LoginRequiredMixin, View):
    """Edit a single council's tracker config."""

    def _get_config(self, council_pk):
        council = get_object_or_404(Council, pk=council_pk)
        cfg, _ = CouncilTrackerConfig.objects.get_or_create(council=council)
        return council, cfg

    def get(self, request, council_pk):
        if not _is_ricd_staff(request.user):
            messages.error(request, "Only RICD staff can edit tracker configuration.")
            return redirect('ui:dashboard')
        council, cfg = self._get_config(council_pk)
        return render(request, 'tracker/config_form.html', {'council': council, 'config': cfg})

    def post(self, request, council_pk):
        if not _is_ricd_staff(request.user):
            messages.error(request, "Only RICD staff can edit tracker configuration.")
            return redirect('ui:dashboard')
        council, cfg = self._get_config(council_pk)
        cfg.council_submission_enabled = request.POST.get('council_submission_enabled') == 'on'
        try:
            cfg.submission_due_day = int(request.POST.get('submission_due_day', 8))
        except ValueError:
            cfg.submission_due_day = 8
        cfg.save()
        messages.success(request, f'Saved tracker config for {council.name}.')
        return redirect('ui:tracker_config_list')


# ===========================================================================
# Monthly Tracker
# ===========================================================================

class MonthlyTrackerListView(LoginRequiredMixin, View):
    """List monthly trackers. Council users see their own; RICD sees all."""

    def get(self, request):
        qs = MonthlyTracker.objects.select_related('council').order_by('-year', '-month', 'council__name')
        if _is_council_user(request.user):
            council = _user_council(request.user)
            if not council:
                messages.warning(request, "Your profile has no council assigned.")
                return redirect('ui:dashboard')
            qs = qs.filter(council=council)
        councils = None
        if _is_ricd_staff(request.user):
            councils = Council.objects.order_by('name')
        return render(request, 'tracker/monthly_list.html', {'trackers': qs, 'councils': councils})


class MonthlyTrackerOpenOrCreateView(LoginRequiredMixin, View):
    """Open the current month's tracker for a council, creating + syncing if absent."""

    def get(self, request, council_pk):
        council = get_object_or_404(Council, pk=council_pk)

        if _is_council_user(request.user):
            if _user_council(request.user) != council:
                raise Http404()

        today = date.today()
        year, month = today.year, today.month
        try:
            yr_arg = int(request.GET.get('year', year))
            mo_arg = int(request.GET.get('month', month))
            if 1 <= mo_arg <= 12:
                year, month = yr_arg, mo_arg
        except (TypeError, ValueError):
            pass

        tracker, created = MonthlyTracker.objects.get_or_create(
            council=council, year=year, month=month
        )
        if created:
            self._sync_entries(tracker)

        return redirect('ui:monthly_tracker_detail', pk=tracker.pk)

    def _sync_entries(self, tracker):
        """Create MonthlyTrackerWorkEntry rows for each WorkStep marked as tracker column."""
        active_projects = _active_projects_for_council(tracker.council)
        works = Work.objects.filter(project__in=active_projects)
        steps = WorkStep.objects.filter(
            work__in=works,
            group_item__is_monthly_tracker_column=True,
            is_active=True,
        ).select_related('work', 'group_item')

        for step in steps:
            MonthlyTrackerWorkEntry.objects.get_or_create(
                tracker=tracker, work_step=step,
                defaults={
                    'actual_completion_date': step.actual_completion_date,
                    'forecast_completion_date': step.forecast_completion_date,
                },
            )


class MonthlyTrackerDetailView(LoginRequiredMixin, View):
    """Grid editor for a single monthly tracker."""

    def _get(self, pk, user):
        tracker = get_object_or_404(MonthlyTracker, pk=pk)
        if _is_council_user(user) and _user_council(user) != tracker.council:
            raise Http404()
        return tracker

    def _can_edit(self, tracker, user):
        if _is_ricd_staff(user):
            return True
        if _is_council_user(user):
            if _user_council(user) != tracker.council:
                return False
            cfg = CouncilTrackerConfig.objects.filter(council=tracker.council).first()
            if not cfg or not cfg.council_submission_enabled:
                return False
            return tracker.status == MonthlyTracker.Status.DRAFT
        return False

    def _build_grid(self, tracker):
        """Rows = Works in active projects; Columns = unique WorkStep definitions.
        Returns columns (list of names) and rows (each with 'cells' aligned to columns).
        """
        entries = (
            tracker.work_entries
            .select_related('work_step__work__project', 'work_step__work__address',
                            'work_step__group_item__step')
            .order_by('work_step__work__project__name',
                      'work_step__work_id', 'work_step__order')
        )
        rows_dict = {}
        column_names = set()
        for e in entries:
            work = e.work_step.work
            col_name = (e.work_step.group_item.step.name
                        if e.work_step.group_item else e.work_step.step_name)
            column_names.add(col_name)
            if work.id not in rows_dict:
                addr = getattr(work, 'address', None)
                if addr:
                    addr_str = f"{addr.street}" + (f" (Lot {addr.lot} {addr.plan})" if getattr(addr, 'lot', None) else "")
                else:
                    addr_str = '(no address)'
                rows_dict[work.id] = {
                    'work': work,
                    'project': work.project,
                    'address': addr_str,
                    'cell_map': {},
                }
            rows_dict[work.id]['cell_map'][col_name] = e

        columns = sorted(column_names)
        # Build cells list aligned to columns
        row_list = []
        for r in sorted(rows_dict.values(),
                        key=lambda r: (r['project'].name if r['project'] else '', r['work'].id)):
            r['cells'] = [r['cell_map'].get(col) for col in columns]
            row_list.append(r)
        return columns, row_list

    def get(self, request, pk):
        tracker = self._get(pk, request.user)
        columns, rows = self._build_grid(tracker)
        return render(request, 'tracker/monthly_detail.html', {
            'tracker': tracker,
            'columns': columns,
            'rows': rows,
            'can_edit': self._can_edit(tracker, request.user),
            'is_ricd': _is_ricd_staff(request.user),
        })

    def post(self, request, pk):
        tracker = self._get(pk, request.user)
        if not self._can_edit(tracker, request.user):
            messages.error(request, 'You do not have permission to edit this tracker.')
            return redirect('ui:monthly_tracker_detail', pk=pk)

        updated = 0
        for e in tracker.work_entries.all():
            actual_field = f'entry_{e.pk}_actual'
            forecast_field = f'entry_{e.pk}_forecast'
            new_actual = date.today() if request.POST.get(actual_field) == 'on' else None

            raw_forecast = request.POST.get(forecast_field) or ''
            new_forecast = None
            if raw_forecast:
                try:
                    from datetime import datetime
                    new_forecast = datetime.strptime(raw_forecast, '%Y-%m-%d').date()
                except ValueError:
                    new_forecast = None

            changed = False
            if e.actual_completion_date != new_actual:
                e.actual_completion_date = new_actual
                changed = True
            if e.forecast_completion_date != new_forecast:
                e.forecast_completion_date = new_forecast
                changed = True
            if changed:
                e.updated_by = request.user
                e.save()
                updated += 1

        if updated:
            messages.success(request, f'Updated {updated} cell(s).')
        return redirect('ui:monthly_tracker_detail', pk=pk)


class MonthlyTrackerSubmitView(LoginRequiredMixin, View):
    """Council submits the tracker."""

    def post(self, request, pk):
        tracker = get_object_or_404(MonthlyTracker, pk=pk)
        if _is_council_user(request.user) and _user_council(request.user) != tracker.council:
            raise Http404()

        cfg = CouncilTrackerConfig.objects.filter(council=tracker.council).first()
        if _is_council_user(request.user) and (not cfg or not cfg.council_submission_enabled):
            messages.error(request, "Submission is not enabled for your council.")
            return redirect('ui:monthly_tracker_detail', pk=pk)

        if tracker.status != MonthlyTracker.Status.DRAFT:
            messages.error(request, f"Cannot submit -- already {tracker.get_status_display()}.")
            return redirect('ui:monthly_tracker_detail', pk=pk)

        tracker.status = MonthlyTracker.Status.SUBMITTED
        tracker.submitted_by = request.user
        tracker.submitted_at = timezone.now()
        tracker.save()
        messages.success(request, 'Monthly tracker submitted.')
        return redirect('ui:monthly_tracker_detail', pk=pk)


class MonthlyTrackerReviewView(LoginRequiredMixin, View):
    """RICD marks a submitted tracker as REVIEWED."""

    def post(self, request, pk):
        if not _is_ricd_staff(request.user):
            messages.error(request, 'Only RICD staff can review trackers.')
            return redirect('ui:monthly_tracker_detail', pk=pk)
        tracker = get_object_or_404(MonthlyTracker, pk=pk)
        if tracker.status != MonthlyTracker.Status.SUBMITTED:
            messages.error(request, 'Only submitted trackers can be reviewed.')
            return redirect('ui:monthly_tracker_detail', pk=pk)
        tracker.status = MonthlyTracker.Status.REVIEWED
        tracker.reviewed_by = request.user
        tracker.reviewed_at = timezone.now()
        tracker.save()
        messages.success(request, 'Tracker marked as reviewed.')
        return redirect('ui:monthly_tracker_detail', pk=pk)


# ===========================================================================
# Quarterly Report
# ===========================================================================

class QuarterlyReportListView(LoginRequiredMixin, View):
    def get(self, request):
        qs = QuarterlyReport.objects.select_related('council').order_by('-year', '-quarter', 'council__name')
        if _is_council_user(request.user):
            council = _user_council(request.user)
            if not council:
                messages.warning(request, "Your profile has no council assigned.")
                return redirect('ui:dashboard')
            qs = qs.filter(council=council)
        councils = None
        if _is_ricd_staff(request.user):
            councils = Council.objects.order_by('name')
        return render(request, 'tracker/quarterly_list.html', {'reports': qs, 'councils': councils})


class QuarterlyReportOpenOrCreateView(LoginRequiredMixin, View):
    """Open the current quarter's report for a council; auto-populate entries."""

    def get(self, request, council_pk):
        council = get_object_or_404(Council, pk=council_pk)
        if _is_council_user(request.user) and _user_council(request.user) != council:
            raise Http404()

        today = date.today()
        try:
            year = int(request.GET.get('year', today.year))
            quarter = int(request.GET.get('quarter', QuarterlyReport.quarter_for_month(today.month)))
            if quarter < 1 or quarter > 4:
                year, quarter = today.year, QuarterlyReport.quarter_for_month(today.month)
        except (TypeError, ValueError):
            year, quarter = today.year, QuarterlyReport.quarter_for_month(today.month)

        report, created = QuarterlyReport.objects.get_or_create(
            council=council, year=year, quarter=quarter
        )
        if created:
            self._sync_entries(report)
        return redirect('ui:quarterly_report_detail', pk=report.pk)

    def _sync_entries(self, report):
        """Pre-create blank entries for Works on Active FSes, using each project's assigned item group."""
        active_fs = FundingSchedule.objects.filter(
            council=report.council,
            status=FundingSchedule.Status.ACTIVE,
        )
        projects = Project.objects.filter(
            funding_schedule__in=active_fs,
            state__in=[Project.State.COMMENCED, Project.State.UNDER_CONSTRUCTION],
        ).exclude(qbuild_delivered=True).select_related('quarterly_report_item_group')

        for project in projects:
            if project.quarterly_report_item_group_id:
                items = list(QuarterlyReportItem.objects.filter(
                    group=project.quarterly_report_item_group,
                    is_active=True,
                ).order_by('order'))
            else:
                items = list(QuarterlyReportItem.objects.filter(
                    is_active=True,
                ).order_by('group__order', 'order'))
            works = Work.objects.filter(project=project)
            for work in works:
                for item in items:
                    QuarterlyReportEntry.objects.get_or_create(
                        report=report, work=work, item=item,
                    )


class QuarterlyReportDetailView(LoginRequiredMixin, View):
    def _get(self, pk, user):
        report = get_object_or_404(QuarterlyReport, pk=pk)
        if _is_council_user(user) and _user_council(user) != report.council:
            raise Http404()
        return report

    def _can_edit(self, report, user):
        if _is_ricd_staff(user):
            return True
        if _is_council_user(user):
            if _user_council(user) != report.council:
                return False
            cfg = CouncilTrackerConfig.objects.filter(council=report.council).first()
            if not cfg or not cfg.council_submission_enabled:
                return False
            return report.status in (QuarterlyReport.Status.DRAFT,
                                     QuarterlyReport.Status.IN_PROGRESS)
        return False

    def _build_grid(self, report):
        """Return FS→Project→Works hierarchy; each project uses its own QR item group columns."""
        # Pre-load all entries for this report into a lookup dict
        entries = report.entries.select_related(
            'work__project__funding_schedule',
            'work__address',
            'item__group',
        ).all()
        cell = {}
        for e in entries:
            cell[(e.work_id, e.item_id)] = e

        active_fs = FundingSchedule.objects.filter(
            council=report.council,
            status=FundingSchedule.Status.ACTIVE,
        ).order_by('schedule_number')

        fs_sections = []
        for fs in active_fs:
            projects = Project.objects.filter(
                funding_schedule=fs,
                state__in=[Project.State.COMMENCED, Project.State.UNDER_CONSTRUCTION],
            ).exclude(qbuild_delivered=True).select_related('quarterly_report_item_group').order_by('name')

            project_sections = []
            for project in projects:
                if project.quarterly_report_item_group_id:
                    items = list(QuarterlyReportItem.objects.filter(
                        group=project.quarterly_report_item_group,
                        is_active=True,
                    ).select_related('group').order_by('order'))
                else:
                    items = list(QuarterlyReportItem.objects.filter(
                        is_active=True,
                    ).select_related('group').order_by('group__order', 'order'))

                works = Work.objects.filter(project=project).order_by('id')
                work_rows = []
                for work in works:
                    addr = getattr(work, 'address', None)
                    if addr:
                        addr_str = f"{addr.street}" + (
                            f" (Lot {addr.lot} {addr.plan})" if getattr(addr, 'lot', None) else ""
                        )
                    else:
                        addr_str = f"Work #{work.id}"
                    work_rows.append({
                        'work': work,
                        'address': addr_str,
                        'cells': [{'item': item, 'entry': cell.get((work.id, item.id))} for item in items],
                    })

                project_sections.append({
                    'project': project,
                    'items': items,
                    'works': work_rows,
                })

            if project_sections:
                fs_sections.append({
                    'fs': fs,
                    'projects': project_sections,
                })

        return {'fs_sections': fs_sections}

    def get(self, request, pk):
        from apps.core.models import SiteSettings
        report = self._get(pk, request.user)
        grid = self._build_grid(report)
        lead = report.council.lead_officer if report.council_id else None
        officer_email = lead.email if lead and lead.email else ''
        officer_name = lead.get_full_name() or lead.username if lead else ''
        reports_email = SiteSettings.get().reports_email
        return render(request, 'tracker/quarterly_detail.html', {
            'report': report,
            'grid': grid,
            'can_edit': self._can_edit(report, request.user),
            'is_ricd': _is_ricd_staff(request.user),
            'reports_email': reports_email,
            'officer_email': officer_email,
            'officer_name': officer_name,
        })

    def post(self, request, pk):
        report = self._get(pk, request.user)
        if not self._can_edit(report, request.user):
            messages.error(request, 'You do not have permission to edit this report.')
            return redirect('ui:quarterly_report_detail', pk=pk)

        updated = 0
        for e in report.entries.select_related('item').all():
            prefix = f'cell_{e.pk}_'
            ft = e.item.field_type
            changed = False

            if ft == 'CHECKBOX':
                val = request.POST.get(prefix + 'bool') == 'on'
                if e.boolean_value != val:
                    e.boolean_value = val
                    changed = True
            elif ft in ('YES_NO', 'YES_NO_NA'):
                raw = request.POST.get(prefix + 'yn', '')
                if raw == 'yes':
                    new_bool, new_na = True, False
                elif raw == 'no':
                    new_bool, new_na = False, False
                elif raw == 'na' and ft == 'YES_NO_NA':
                    new_bool, new_na = None, True
                else:
                    new_bool, new_na = None, False
                if e.boolean_value != new_bool or e.is_na != new_na:
                    e.boolean_value = new_bool
                    e.is_na = new_na
                    changed = True
            elif ft in ('DATE', 'DATE_NA'):
                raw = request.POST.get(prefix + 'date', '')
                na = request.POST.get(prefix + 'na') == 'on' if ft == 'DATE_NA' else False
                new_date = None
                if raw and not na:
                    try:
                        from datetime import datetime
                        new_date = datetime.strptime(raw, '%Y-%m-%d').date()
                    except ValueError:
                        new_date = None
                if e.date_value != new_date or e.is_na != na:
                    e.date_value = new_date
                    e.is_na = na
                    changed = True
            elif ft in ('NUMBER', 'CURRENCY'):
                raw = request.POST.get(prefix + 'num', '')
                from decimal import Decimal, InvalidOperation
                try:
                    new_num = Decimal(raw) if raw else None
                except InvalidOperation:
                    new_num = None
                if e.number_value != new_num:
                    e.number_value = new_num
                    changed = True
            else:  # TEXT
                raw = request.POST.get(prefix + 'text', '')
                if e.text_value != raw:
                    e.text_value = raw
                    changed = True

            if changed:
                e.updated_by = request.user
                e.save()
                updated += 1

        # Save report-level fields (adverse matters, declaration)
        report_changed = False
        if request.POST.get('no_adverse') == 'on':
            new_adverse = ''
        else:
            new_adverse = request.POST.get('adverse_matters', report.adverse_matters)
        if report.adverse_matters != new_adverse:
            report.adverse_matters = new_adverse
            report_changed = True

        for field in ('declaration_officer_name', 'declaration_officer_position'):
            val = request.POST.get(field, '')
            if getattr(report, field) != val:
                setattr(report, field, val)
                report_changed = True

        decl_date_raw = request.POST.get('declaration_date', '')
        if decl_date_raw:
            try:
                from datetime import datetime
                new_decl_date = datetime.strptime(decl_date_raw, '%Y-%m-%d').date()
            except ValueError:
                new_decl_date = report.declaration_date
        else:
            new_decl_date = None
        if report.declaration_date != new_decl_date:
            report.declaration_date = new_decl_date
            report_changed = True

        if report.status == QuarterlyReport.Status.DRAFT and (updated or report_changed):
            report.status = QuarterlyReport.Status.IN_PROGRESS
            report_changed = True

        if report_changed:
            report.save()

        if updated or report_changed:
            messages.success(request, f'Report saved.')
        return redirect('ui:quarterly_report_detail', pk=pk)


class QuarterlyReportSubmitView(LoginRequiredMixin, View):
    def post(self, request, pk):
        report = get_object_or_404(QuarterlyReport, pk=pk)
        if _is_council_user(request.user) and _user_council(request.user) != report.council:
            raise Http404()
        cfg = CouncilTrackerConfig.objects.filter(council=report.council).first()
        if _is_council_user(request.user) and (not cfg or not cfg.council_submission_enabled):
            messages.error(request, "Submission is not enabled for your council.")
            return redirect('ui:quarterly_report_detail', pk=pk)
        if report.status not in (QuarterlyReport.Status.DRAFT,
                                 QuarterlyReport.Status.IN_PROGRESS):
            messages.error(request, f"Cannot submit -- already {report.get_status_display()}.")
            return redirect('ui:quarterly_report_detail', pk=pk)
        report.status = QuarterlyReport.Status.SUBMITTED
        report.submitted_by = request.user
        report.submitted_at = timezone.now()
        report.save()
        messages.success(request, 'Quarterly report submitted.')
        return redirect('ui:quarterly_report_detail', pk=pk)


class QuarterlyReportApproveView(LoginRequiredMixin, View):
    """RICD MANAGER/DIRECTOR approves the quarterly report."""

    def post(self, request, pk):
        if _role(request.user) not in MANAGER_ROLES and not request.user.is_superuser:
            messages.error(request, 'Only Managers/Directors can approve quarterly reports.')
            return redirect('ui:quarterly_report_detail', pk=pk)
        report = get_object_or_404(QuarterlyReport, pk=pk)
        if report.status != QuarterlyReport.Status.SUBMITTED:
            messages.error(request, 'Only submitted reports can be approved.')
            return redirect('ui:quarterly_report_detail', pk=pk)
        report.status = QuarterlyReport.Status.APPROVED
        report.approved_by = request.user
        report.approved_at = timezone.now()
        report.save()
        messages.success(request, 'Quarterly report approved.')
        return redirect('ui:quarterly_report_detail', pk=pk)


# ===========================================================================
# Monthly Tracker — Excel export / import
# ===========================================================================

class MonthlyTrackerExportView(LoginRequiredMixin, View):
    """Download the tracker as a pre-filled .xlsx for offline council completion.

    Column layout (1-indexed):
      A: _entry_pk  — system ID, do not edit
      B: Project
      C: Work / Address
      D: Step
      E: Actual Date       ← council fills in
      F: Forecast Date     ← council fills in
      G: Notes             ← council fills in
    """

    def get(self, request, pk):
        if not _is_ricd_staff(request.user):
            messages.error(request, 'Only RICD staff can export trackers.')
            return redirect('ui:monthly_tracker_detail', pk=pk)

        tracker = get_object_or_404(MonthlyTracker, pk=pk)
        entries = list(
            tracker.work_entries
            .select_related(
                'work_step__work__project',
                'work_step__work__address',
                'work_step__group_item__step',
            )
            .order_by(
                'work_step__work__project__name',
                'work_step__work_id',
                'work_step__order',
            )
        )

        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{tracker.year}-{tracker.month:02d}"

        grey = PatternFill('solid', fgColor='DDDDDD')
        yellow = PatternFill('solid', fgColor='FFFACD')
        locked = PatternFill('solid', fgColor='F5F5F5')

        ws['A1'] = (
            f"Monthly Tracker — {tracker.council.name} "
            f"— {tracker.year}-{tracker.month:02d}"
        )
        ws['A1'].font = Font(bold=True, size=13)
        ws.merge_cells('A1:G1')

        ws['A2'] = (
            "Fill in columns E (Actual Date), F (Forecast Date) and G (Notes) only. "
            "Do not add or remove rows. Return completed file to RICD."
        )
        ws['A2'].font = Font(italic=True, color='666666', size=10)
        ws.merge_cells('A2:G2')

        headers = [
            '_entry_pk', 'Project', 'Work / Address', 'Step',
            'Actual Date (dd/mm/yyyy)', 'Forecast Date (dd/mm/yyyy)', 'Notes',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = grey

        for i, entry in enumerate(entries, start=5):
            ws_step = entry.work_step
            work = ws_step.work
            project = work.project
            addr = getattr(work, 'address', None)
            if addr:
                addr_str = addr.street + (
                    f" (Lot {addr.lot} {addr.plan})" if getattr(addr, 'lot', None) else ""
                )
            else:
                addr_str = f"Work #{work.id}"
            step_name = (
                ws_step.group_item.step.name if ws_step.group_item_id else ws_step.step_name
            )

            for col, val in enumerate(
                [entry.pk, project.name if project else '', addr_str, step_name], 1
            ):
                ws.cell(row=i, column=col, value=val).fill = locked

            actual = ws.cell(row=i, column=5)
            if entry.actual_completion_date:
                actual.value = entry.actual_completion_date
                actual.number_format = 'DD/MM/YYYY'
            actual.fill = yellow

            forecast = ws.cell(row=i, column=6)
            if entry.forecast_completion_date:
                forecast.value = entry.forecast_completion_date
                forecast.number_format = 'DD/MM/YYYY'
            forecast.fill = yellow

            ws.cell(row=i, column=7, value=entry.notes or '').fill = yellow

        for letter, width in zip('ABCDEFG', [9, 26, 32, 22, 24, 24, 36]):
            ws.column_dimensions[letter].width = width

        fname = (
            f"monthly-tracker"
            f"-{tracker.council.name.lower().replace(' ', '-')}"
            f"-{tracker.year}-{tracker.month:02d}.xlsx"
        )
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response


class MonthlyTrackerImportView(LoginRequiredMixin, View):
    """Accept a council-completed .xlsx and write the data back into the tracker.

    Matches rows by the hidden _entry_pk in column A.
    Calls entry.save() for each changed row so the WorkStep sync signal fires.
    Marks the tracker SUBMITTED after a successful import.
    """

    def post(self, request, pk):
        if not _is_ricd_staff(request.user):
            messages.error(request, 'Only RICD staff can import trackers.')
            return redirect('ui:monthly_tracker_detail', pk=pk)

        tracker = get_object_or_404(MonthlyTracker, pk=pk)
        xlsx_file = request.FILES.get('tracker_file')
        if not xlsx_file:
            messages.error(request, 'No file selected.')
            return redirect('ui:monthly_tracker_detail', pk=pk)

        import openpyxl
        from datetime import datetime as dt

        try:
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        except Exception:
            messages.error(request, 'Could not open the file — ensure it is an .xlsx file.')
            return redirect('ui:monthly_tracker_detail', pk=pk)

        ws = wb.active
        entry_map = {e.pk: e for e in tracker.work_entries.all()}
        updated = skipped = 0
        errors = []

        def _parse_date(val):
            if val is None:
                return None
            if hasattr(val, 'year'):
                return val.date() if hasattr(val, 'date') else val
            s = str(val).strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
                try:
                    return dt.strptime(s, fmt).date()
                except ValueError:
                    pass
            return None

        for row in ws.iter_rows(min_row=5, values_only=True):
            if all(v is None for v in row):
                continue
            try:
                entry_pk = int(row[0])
            except (TypeError, ValueError):
                skipped += 1
                continue

            entry = entry_map.get(entry_pk)
            if entry is None:
                errors.append(f"Entry #{entry_pk} not found in this tracker — row skipped.")
                skipped += 1
                continue

            new_actual = _parse_date(row[4])
            new_forecast = _parse_date(row[5])
            new_notes = str(row[6]).strip() if row[6] else ''

            changed = False
            if entry.actual_completion_date != new_actual:
                entry.actual_completion_date = new_actual
                changed = True
            if entry.forecast_completion_date != new_forecast:
                entry.forecast_completion_date = new_forecast
                changed = True
            if entry.notes != new_notes:
                entry.notes = new_notes
                changed = True
            if changed:
                entry.updated_by = request.user
                entry.save()
                updated += 1

        for msg in errors[:5]:
            messages.warning(request, msg)

        if updated:
            if tracker.status == MonthlyTracker.Status.DRAFT:
                tracker.status = MonthlyTracker.Status.SUBMITTED
                tracker.submitted_by = request.user
                tracker.submitted_at = timezone.now()
                tracker.save()
            messages.success(
                request,
                f'Import complete: {updated} row(s) updated. Tracker marked as Submitted.',
            )
        else:
            messages.info(
                request,
                f'No changes detected in the file ({skipped} row(s) skipped).',
            )
