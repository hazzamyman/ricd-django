from django.core.management.base import BaseCommand
from django.utils import timezone
from decimal import Decimal
from ricd.models import Project, Council, Program, FundingSchedule, Work
import openpyxl
from openpyxl import load_workbook
import os

class Command(BaseCommand):
    help = 'Import master data from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Path to the master data Excel file',
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='Sheet1',
            help='Sheet name to import from (default: Sheet1)',
        )

    def handle(self, *args, **options):
        file_path = options['file']
        sheet_name = options['sheet']

        if not file_path:
            self.stdout.write(self.style.ERROR('Please provide --file path'))
            return

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'File does not exist: {file_path}'))
            return

        # Column mapping from Excel headers to model fields
        column_mapping = {
            'RICD Rep as at 7 Feb 2025': 'ricd_rep_date',
            'Commencement (LOA) Date Forecast': 'commencement_loa_forecast',
            'Commencement (LOA) Date Actual': 'commencement_loa_actual',
            'Date physically commenced on site': 'date_physically_commenced',
            'Estimated Completion': 'estimated_completion',
            'Actual Completion Date': 'actual_completion',
            'Handover Forecast': 'handover_forecast',
            'Handover Actual': 'handover_actual',
            'Month Tenancy Agreement Signed': 'month_tenancy_signed',
            'SAP Master Project': 'sap_master_project',
            'SAP Project': 'sap_project',
            'Project Manager': 'project_manager',
            'BE': 'be',
            'BU': 'bu',
            'RU': 'ru',
            'CLI No': 'cli_no',
            'Street Address': 'street_address',
            'Previous Address': 'previous_address',
            'Suburb Town': 'suburb_town',
            'LGA': 'lga',
            'RPD': '',  # Skip or add field if needed
            'Project Yield': '',  # Skip
            'Floor No Eg 1': 'floor_no',
            '40 Yr Lease Site Reference': '',  # Skip
            'Type of Land Tenure': 'type_of_land_tenure',
            'Month Secure Tenure Executed': 'month_secure_tenure_executed',
            'Lease Commencement Date': '',  # Skip
            'Planned Lease Commencement Lease Summary': '',  # Skip
            'Drawing No': 'drawing_no',
            'Umbrella Agreement Schedule': '',  # Skip
            'State Financial Approval': '',  # Skip
            'Type of Approval': '',  # Skip
            'Program': 'program_field',  # Conflicts with foreign key, use carefully
            'Funding Agreement': 'funding_agreement',
            'Quickstarts': 'quickstarts',
            'Homes for Queenslanders completions': 'homes_for_queenslanders_completions',
            'Land Status': 'land_status',
            'Usage Type': 'usage_type',
            'Initial Date of CAA': 'initial_date_of_caa',
            'Prog Year (FY Commenced)': 'program_year',
            'Package': 'package',
            'House Type': 'house_type',
            'Contractor': 'contractor',
            "Contractor's Address": 'contractor_address',
            'Construction Costs': 'construction_costs',
            'Floor Area': 'floor_area',
            'Total Outputs': '',  # Skip, could be calculations
            'Output Type': '',  # Skip
            'Fully Adaptable Dwellings': '',  # Skip
            'Semi Adaptable Dwellings': '',  # Skip
            'FLOOR (Concrete Slab/Timber Frame/Steel Frame)': 'floor_method',
            'FRAME (Timber Frame/Steel Frame/Block/FC Panel)': 'frame_method',
            'EXTERNAL WALL (Timber/Sheeting/Block/Brick)': 'external_wall_method',
            'ROOF (Metal Sheeting /Tiles/Galv.Sheeting/Colourbond)': 'roof_method',
            'CAR ACCOM. (Carport/Garage/Under House)': 'car_accommodation',
            'ADDITIONAL FACILITIES: WC/BATHROOM': 'additional_facilities',
            'Version 120 Budget': 'version_120_budget',
            'Forecast Final Cost': 'forecast_final_cost',
            'Final Cost': 'final_cost',
            'Costs Finalised': 'costs_finalised',
            'State Electorate': 'state_electorate',
            'QHIGI Region': 'qhigi_region',
            'Federal Electorate': 'federal_electorate',
            'Notes / Comments': 'notes_comments',
            'Comments 1': 'comments_1',
            'Comments 2': 'comments_2',
            'CIMS Number / Reside Act Ref': 'cims_number',
            'Commence (LOA)': 'commencement_loa_forecast',  # Same as earlier
            # Skip totals and other calculated fields
        }

        # Load workbook
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            self.stdout.write(self.style.ERROR(f'Sheet "{sheet_name}" not found. Available: {wb.sheetnames}'))
            return

        ws = wb[sheet_name]

        # Get header row (assume first row is headers)
        headers = [cell.value for cell in ws[1]]

        # Create mappings dictionary
        row_dicts = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for header, value in zip(headers, row):
                if header and str(header).strip() in column_mapping:
                    field_name = column_mapping[str(header).strip()]
                    if field_name:  # Skip empty mappings
                        row_dict[field_name] = value
            if row_dict:
                row_dicts.append(row_dict)

        # Default council and program (need to create or get)
        council, _ = Council.objects.get_or_create(name='Default Council')  # Adjust as needed
        program, _ = Program.objects.get_or_create(name='Default Program')   # Adjust as needed
        funding_schedule, _ = FundingSchedule.objects.get_or_create(
            council=council,
            funding_schedule_number=1,  # Generic
            funding_amount=Decimal('0')  # Will be updated
        )

        projects_created = 0
        for row in row_dicts:
            # Create Project instance
            project_data = {
                'name': f"Project {row.get('sap_project', '')}",  # Or derive from other fields
                'lga': row.get('lga'),
                'federal_electorate': row.get('federal_electorate'),
                'state_electorate': row.get('state_electorate'),
                'qhigi_region': row.get('qhigi_region'),
                'suburb_town': row.get('suburb_town'),
                'street_address': row.get('street_address'),
                'previous_address': row.get('previous_address'),
                'contractor': row.get('contractor'),
                'contractor_address': row.get('contractor_address'),
                'construction_costs': self.clean_decimal(row.get('construction_costs')),
                'floor_area': self.clean_decimal(row.get('floor_area')),
                'house_type': row.get('house_type'),
                'program_year': row.get('program_year') or '2023-24',  # Default
                'commitments': self.clean_decimal(row.get('commitments')),
                'final_cost': self.clean_decimal(row.get('final_cost')),
                'forecast_final_cost': self.clean_decimal(row.get('forecast_final_cost')),
                'costs_finalised': row.get('costs_finalised') == 'Yes',  # Adjust based on data
                'handover_forecast': self.clean_date(row.get('handover_forecast')),
                'handover_actual': self.clean_date(row.get('handover_actual')),
                'commencement_loa_forecast': self.clean_date(row.get('commencement_loa_forecast')),
                'commencement_loa_actual': self.clean_date(row.get('commencement_loa_actual')),
                'date_physically_commenced': self.clean_date(row.get('date_physically_commenced')),
                'estimated_completion': self.clean_date(row.get('estimated_completion')),
                'actual_completion': self.clean_date(row.get('actual_completion')),
                'floor_no': row.get('floor_no'),
                'type_of_land_tenure': row.get('type_of_land_tenure'),
                'month_secure_tenure_executed': self.clean_date(row.get('month_secure_tenure_executed')),
                'drawing_no': row.get('drawing_no'),
                'program': row.get('program_field'),
                'funding_agreement': row.get('funding_agreement'),
                'package': row.get('package'),
                'quickstarts': row.get('quickstarts'),
                'land_status': row.get('land_status'),
                'usage_type': row.get('usage_type'),
                'initial_date_of_caa': self.clean_date(row.get('initial_date_of_caa')),
                'version_120_budget': self.clean_decimal(row.get('version_120_budget')),
                'sap_master_project': row.get('sap_master_project'),
                'sap_project': row.get('sap_project'),
                'project_manager': row.get('project_manager'),
                'cli_no': row.get('cli_no'),
                'be': row.get('be'),
                'bu': row.get('bu'),
                'ru': row.get('ru'),
                'notes_comments': row.get('notes_comments'),
                'comments_1': row.get('comments_1'),
                'comments_2': row.get('comments_2'),
                'cims_number': row.get('cims_number'),
                # Set foreign keys
                'council': council,
                'program': program,
                'funding_schedule': funding_schedule,
            }

            # Create Work instance if needed (related to floors, bathrooms, etc.)
            work_data = {
                'floor_method': row.get('floor_method'),
                'frame_method': row.get('frame_method'),
                'external_wall_method': row.get('external_wall_method'),
                'roof_method': row.get('roof_method'),
                'car_accommodation': row.get('car_accommodation'),
                'additional_facilities': row.get('additional_facilities'),
            }

            try:
                # Clean data and set
                for k, v in project_data.items():
                    if v is not None and k.endswith('_amount') or k.endswith('_cost'):
                        project_data[k] = self.clean_decimal(v)
                project, created = Project.objects.update_or_create(
                    sap_project=row.get('sap_project'),
                    defaults=project_data
                )

                if created:
                    projects_created += 1
                    self.stdout.write(f"Created project: {project.name}")

                    # Create related Work if construction data present
                    if any(row.get(key) for key in ['floor_method', 'bedrooms', 'bathrooms']):
                        Work.objects.create(
                            project=project,
                            work_type='construction',
                            output_type='detached_house',  # Default, adjust
                            bedrooms=row.get('bedrooms', 0),
                            bathrooms=row.get('bathrooms', 0),
                            kitchens=row.get('kitchens', 0),
                            floor_method=row.get('floor_method'),
                            frame_method=row.get('frame_method'),
                            external_wall_method=row.get('external_wall_method'),
                            roof_method=row.get('roof_method'),
                            car_accommodation=row.get('car_accommodation'),
                            additional_facilities=row.get('additional_facilities'),
                        )

            except Exception as e:
                self.stdout.write(self.style.ERROR(f"Error creating project: {e}, data: {row}"))

        self.stdout.write(self.style.SUCCESS(f'Successfully imported {projects_created} projects'))

    def clean_decimal(self, value):
        """Convert value to Decimal or return None"""
        if value is None:
            return None
        try:
            return Decimal(str(value).replace('$', '').replace(',', ''))
        except:
            return None

    def clean_date(self, value):
        """Convert various date formats to date object"""
        if not value:
            return None
        if isinstance(value, str):
            # Handle 'XX-XXXX' or similar date strings, adjust parsing as needed
            try:
                # Assume format like '01/01/2023', adjust if different
                day, month, year = value.split('/')
                return timezone.datetime(int(year), int(month), int(day)).date()
            except:
                return None
        elif hasattr(value, 'date'):
            return value.date()
        return value

    def clean_boolean(self, value):
        """Convert value to boolean"""
        if value is None:
            return None
        return str(value).lower() in ['yes', 'true', '1', 'y']